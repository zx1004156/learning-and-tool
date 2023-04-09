import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font
import tkinter as tk
from tkinter import filedialog
import os
import time
import datetime
today = datetime.date.today()


file_path = filedialog.askopenfilename()
if not file_path:
    os._exit(0)
mywb = load_workbook(
    file_path)

start = time.time()

mywb.create_sheet("三民店(SM)")
mywb.create_sheet("土城店(TC)")
mywb.create_sheet("中清店(CS)")
mywb.create_sheet("北投店(BE)")
#mywb.create_sheet("北門店(BM)")
mywb.create_sheet("北港店(BK)")
mywb.create_sheet("板橋店(PC)")
mywb.create_sheet("苗栗店(ML)")
mywb.create_sheet("新仁店(XJ)")

sheet = mywb["工作表1"]
sheet1 = mywb["三民店(SM)"]
sheet2 = mywb["土城店(TC)"]
sheet3 = mywb["中清店(CS)"]
sheet4 = mywb["北投店(BE)"]
#sheet5 = mywb["北門店(BM)"]
sheet6 = mywb["北港店(BK)"]
sheet7 = mywb["板橋店(PC)"]
sheet8 = mywb["苗栗店(ML)"]
sheet9 = mywb["新仁店(XJ)"]


sheet1['A1'] = '店號'
sheet1['B1'] = '庫別'
sheet1['C1'] = '店名'
sheet1['D1'] = '條碼'
sheet1['E1'] = '品號'
sheet1['F1'] = '品名'
sheet1['G1'] = '規格'
sheet1['H1'] = '單位'
sheet1['I1'] = '數量'

sheet2['A1'] = '店號'
sheet2['B1'] = '庫別'
sheet2['C1'] = '店名'
sheet2['D1'] = '條碼'
sheet2['E1'] = '品號'
sheet2['F1'] = '品名'
sheet2['G1'] = '規格'
sheet2['H1'] = '單位'
sheet2['I1'] = '數量'

sheet3['A1'] = '店號'
sheet3['B1'] = '庫別'
sheet3['C1'] = '店名'
sheet3['D1'] = '條碼'
sheet3['E1'] = '品號'
sheet3['F1'] = '品名'
sheet3['G1'] = '規格'
sheet3['H1'] = '單位'
sheet3['I1'] = '數量'

sheet4['A1'] = '店號'
sheet4['B1'] = '庫別'
sheet4['C1'] = '店名'
sheet4['D1'] = '條碼'
sheet4['E1'] = '品號'
sheet4['F1'] = '品名'
sheet4['G1'] = '規格'
sheet4['H1'] = '單位'
sheet4['I1'] = '數量'

'''
sheet5['A1'] = '店號'
sheet5['B1'] = '庫別'
sheet5['C1'] = '店名'
sheet5['D1'] = '條碼'
sheet5['E1'] = '品號'
sheet5['F1'] = '品名'
sheet5['G1'] = '規格'
sheet5['H1'] = '單位'
sheet5['I1'] = '數量'
'''
sheet6['A1'] = '店號'
sheet6['B1'] = '庫別'
sheet6['C1'] = '店名'
sheet6['D1'] = '條碼'
sheet6['E1'] = '品號'
sheet6['F1'] = '品名'
sheet6['G1'] = '規格'
sheet6['H1'] = '單位'
sheet6['I1'] = '數量'

sheet7['A1'] = '店號'
sheet7['B1'] = '庫別'
sheet7['C1'] = '店名'
sheet7['D1'] = '條碼'
sheet7['E1'] = '品號'
sheet7['F1'] = '品名'
sheet7['G1'] = '規格'
sheet7['H1'] = '單位'
sheet7['I1'] = '數量'

sheet8['A1'] = '店號'
sheet8['B1'] = '庫別'
sheet8['C1'] = '店名'
sheet8['D1'] = '條碼'
sheet8['E1'] = '品號'
sheet8['F1'] = '品名'
sheet8['G1'] = '規格'
sheet8['H1'] = '單位'
sheet8['I1'] = '數量'

sheet9['A1'] = '店號'
sheet9['B1'] = '庫別'
sheet9['C1'] = '店名'
sheet9['D1'] = '條碼'
sheet9['E1'] = '品號'
sheet9['F1'] = '品名'
sheet9['G1'] = '規格'
sheet9['H1'] = '單位'
sheet9['I1'] = '數量'

a = []
b = []
c = []
d = []
e = []
f = []
g = []
h = []
i = []


for n in range(1, sheet.max_row+1):
    a.append(sheet.cell(row = n+1, column = 1).value)
    b.append(sheet.cell(row = n+1, column = 2).value)
    c.append(sheet.cell(row = n+1, column = 3).value)
    d.append(sheet.cell(row = n+1, column = 4).value)
    e.append(sheet.cell(row = n+1, column = 5).value)
    f.append(sheet.cell(row = n+1, column = 6).value)
    g.append(sheet.cell(row = n+1, column = 7).value)
    h.append(sheet.cell(row = n+1, column = 8).value)
    i.append(sheet.cell(row = n+1, column = 9).value)

    s1 = 2
    s2 = 2
    s3 = 2
    s4 = 2
    s5 = 2
    s6 = 2
    s7 = 2
    s8 = 2
    s9 = 2

for t1 in range(len(c)):
    if c[t1] == "三民店(SM)":
        sheet1['A'+str(s1)] = a[t1]
        sheet1['B'+str(s1)] = b[t1]
        sheet1['C'+str(s1)] = c[t1]
        sheet1['D'+str(s1)] = d[t1]
        sheet1['E'+str(s1)] = e[t1]
        sheet1['F'+str(s1)] = f[t1]
        sheet1['G'+str(s1)] = g[t1]
        sheet1['H'+str(s1)] = h[t1]
        sheet1['I'+str(s1)] = i[t1]
        s1+=1
    elif c[t1] == "土城店TC)":
        sheet2['A'+str(s2)] = a[t1]
        sheet2['B'+str(s2)] = b[t1]
        sheet2['C'+str(s2)] = c[t1]
        sheet2['D'+str(s2)] = d[t1]
        sheet2['E'+str(s2)] = e[t1]
        sheet2['F'+str(s2)] = f[t1]
        sheet2['G'+str(s2)] = g[t1]
        sheet2['H'+str(s2)] = h[t1]
        sheet2['I'+str(s2)] = i[t1]
        s2+=1
    
    elif c[t1] == "中清店(CS)":
        sheet3['A'+str(s3)] = a[t1]
        sheet3['B'+str(s3)] = b[t1]
        sheet3['C'+str(s3)] = c[t1]
        sheet3['D'+str(s3)] = d[t1]
        sheet3['E'+str(s3)] = e[t1]
        sheet3['F'+str(s3)] = f[t1]
        sheet3['G'+str(s3)] = g[t1]
        sheet3['H'+str(s3)] = h[t1]
        sheet3['I'+str(s3)] = i[t1]
        s3+=1

    elif c[t1] == "北投店(BE)":
        sheet4['A'+str(s4)] = a[t1]
        sheet4['B'+str(s4)] = b[t1]
        sheet4['C'+str(s4)] = c[t1]
        sheet4['D'+str(s4)] = d[t1]
        sheet4['E'+str(s4)] = e[t1]
        sheet4['F'+str(s4)] = f[t1]
        sheet4['G'+str(s4)] = g[t1]
        sheet4['H'+str(s4)] = h[t1]
        sheet4['I'+str(s4)] = i[t1]
        s4+=1

    elif c[t1] == "北港店 (BK)":
        sheet6['A'+str(s6)] = a[t1]
        sheet6['B'+str(s6)] = b[t1]
        sheet6['C'+str(s6)] = c[t1]
        sheet6['D'+str(s6)] = d[t1]
        sheet6['E'+str(s6)] = e[t1]
        sheet6['F'+str(s6)] = f[t1]
        sheet6['G'+str(s6)] = g[t1]
        sheet6['H'+str(s6)] = h[t1]
        sheet6['I'+str(s6)] = i[t1]
        s6+=1

    elif c[t1] == "板橋店(PC)":
        sheet7['A'+str(s7)] = a[t1]
        sheet7['B'+str(s7)] = b[t1]
        sheet7['C'+str(s7)] = c[t1]
        sheet7['D'+str(s7)] = d[t1]
        sheet7['E'+str(s7)] = e[t1]
        sheet7['F'+str(s7)] = f[t1]
        sheet7['G'+str(s7)] = g[t1]
        sheet7['H'+str(s7)] = h[t1]
        sheet7['I'+str(s7)] = i[t1]
        s7+=1

    elif c[t1] == "苗栗店(ML)":
        sheet8['A'+str(s8)] = a[t1]
        sheet8['B'+str(s8)] = b[t1]
        sheet8['C'+str(s8)] = c[t1]
        sheet8['D'+str(s8)] = d[t1]
        sheet8['E'+str(s8)] = e[t1]
        sheet8['F'+str(s8)] = f[t1]
        sheet8['G'+str(s8)] = g[t1]
        sheet8['H'+str(s8)] = h[t1]
        sheet8['I'+str(s8)] = i[t1]
        s8+=1

    elif c[t1] == "新仁店 (XJ)":
        sheet9['A'+str(s9)] = a[t1]
        sheet9['B'+str(s9)] = b[t1]
        sheet9['C'+str(s9)] = c[t1]
        sheet9['D'+str(s9)] = d[t1]
        sheet9['E'+str(s9)] = e[t1]
        sheet9['F'+str(s9)] = f[t1]
        sheet9['G'+str(s9)] = g[t1]
        sheet9['H'+str(s9)] = h[t1]
        sheet9['I'+str(s9)] = i[t1]
        s9+=1



del mywb["工作表2"]
del mywb["工作表3"]        
mywb.save(str(today)+ "保健店消耗量(已分類).xlsx")  # 另存新檔


end = time.time()
print("花費時間"+ str(end-start))
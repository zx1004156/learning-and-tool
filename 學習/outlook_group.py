from openpyxl import load_workbook
import win32com.client

# 讀取 Excel 表格
wb = load_workbook(r'C:\Users\user\Downloads\sample_data3.xlsx')
ws = wb.worksheets[0]
group_name = ws['A1'].value
member1 = ws['A2'].value
member2 = ws['A3'].value
member3 = ws['A4'].value

# 創建 Outlook 物件
outlook = win32com.client.Dispatch("Outlook.Application")

# 創建聯繫人群組
group = outlook.CreateItem(10)
group.Subject = group_name

# 添加聯繫人
group.AddMember(member1)
group.AddMember(member2)
group.AddMember(member3)

# 保存群組
group.Save()
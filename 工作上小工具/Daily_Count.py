from openpyxl import load_workbook
import tkinter as tk
from tkinter import filedialog
import os


file_path = filedialog.askopenfilename()
if not file_path:
    os._exit(0)
mywb = load_workbook(
    file_path)

mywb.create_sheet("每日")  # 創建一個叫TEST的表
sheet = mywb['Sheet1']  # 抓取名叫"詳細"的表
sheet2 = mywb['每日']
# print(sheet['D3'].value[0:10])  # <class 'str'>  取出前十個數 日期

cellRange = sheet['D3':'D1000']  # 範圍刻意設大一點

arr = [0] * 32  # 串列設大小31日
t = 3
t2 = -1
t3 = 1

sheet['D3']
year = sheet['D3'].value[0:4]
month = sheet['D3'].value[5:7]

for i in cellRange:
    for j in i:
        if (j.value != None):
            # print(j.value[5:7] + j.value[8:10])日期
            for k in range(10):  # 1號到9號
                if j.value[5:7] + j.value[8:10] == month+"0"+str(k):
                    #print(j.value[5:7] + j.value[8:10])
                    arr[k] += sheet['G' + str(t)].value
                    t += 1
            for l in range(10, 32):  # 10號到31號
                if j.value[5:7] + j.value[8:10] == month+str(l):
                    arr[l] += sheet['G' + str(t)].value
                    t += 1

# 以下為寫入excel
sheet2['A1'] = '日期'
sheet2['B1'] = '幣別'
sheet2['C1'] = '金額'

coin_kind = sheet['F3']

for m in arr:
    t2 += 1  # 日期計數
    if m != 0:
        t3 += 1  # excel欄位計數
        sheet2['A' + str(t3)] = year +'/'+month+'/' + str(t2)
        sheet2['B' + str(t3)] = coin_kind.value
        sheet2['C' + str(t3)] = arr[t2]
        print(str(m)+' '+str(t2)+' '+str(t3))


mywb.save(r'C:\Users\user\Downloads\python計算.xlsx')  # 另存新檔
